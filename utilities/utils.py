"""UTILITIES ES UN CONCEPTO MUY SIMILAR A LA BASE CLASS, ES PARA RE UTILIZAR METODOS COMMUNES
EN UN FRAMEWORK PONERLO EN LAS UTILITIES.
SI NO TIENEN DRIVER VAN ACA, SI TIENEN DRIVER VAN A LA BASE"""
import inspect
import logging
import softest
from openpyxl import Workbook, load_workbook
import csv


class Utils(softest.TestCase):

    def assert_list_item_text(self, list_to_check, value_to_verify):
        for flight in list_to_check:
            print("The text is: " + flight.text)
            self.soft_assert(self.assertTrue, (value_to_verify in flight.text))
            if value_to_verify in flight.text:
                print('Assert passed')
            else:
                print('Assert NOT passed')
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        # Gets the name of the class / method from where this method is called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler('automation.log', mode='a')
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s:  %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]

        """Obtener maximo de columnas y filas"""
        row_count = sh.max_row
        column_count = sh.max_column

        """Imprimir los valores de filas y columnas"""
        for i in range(2, row_count + 1):
            row = []
            for j in range(1, column_count + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(file_name):
        # Create an empty list
        datalist = []
        # Open CSV file
        csvdata = open(file_name, "r")
        # Create CSV reader
        reader = csv.reader(csvdata)
        # Skip header
        next(reader)
        # Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist





